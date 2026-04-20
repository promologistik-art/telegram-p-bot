import os
import logging
from datetime import datetime
from io import BytesIO
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes
from sqlalchemy import select, func, delete
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from config import Config
from database import AsyncSessionLocal
from models import User, Project, PostQueue
from backup import BackupService
from .utils import is_admin, get_user_projects_count

logger = logging.getLogger(__name__)


async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await is_admin(update.effective_user.id):
        await update.message.reply_text("❌ Нет доступа")
        return
    
    keyboard = [
        [InlineKeyboardButton("👥 Пользователи", callback_data="admin_users_list")],
        [InlineKeyboardButton("💾 Создать бэкап", callback_data="admin_backup_create")],
        [InlineKeyboardButton("📦 Список бэкапов", callback_data="admin_backup_list")],
        [InlineKeyboardButton("📊 Экспорт в Excel", callback_data="admin_export")],
        [InlineKeyboardButton("🔍 Диагностика", callback_data="admin_diagnose")],
        [InlineKeyboardButton("🧹 Очистить очередь", callback_data="admin_clear_queue")],
        [InlineKeyboardButton("🗑️ Очистить failed", callback_data="admin_clear_failed")],
    ]
    
    if update.callback_query:
        await update.callback_query.edit_message_text(
            "👑 <b>Админ-панель</b>\n\nВыберите действие:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="HTML"
        )
    else:
        await update.message.reply_text(
            "👑 <b>Админ-панель</b>\n\nВыберите действие:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="HTML"
        )


async def admin_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    if not await is_admin(update.effective_user.id):
        await query.edit_message_text("❌ Нет доступа")
        return
    
    action = query.data
    
    if action == "admin_users_list":
        await show_admin_users(query)
    elif action == "admin_backup_create":
        await create_backup_admin(query)
    elif action == "admin_backup_list":
        await list_backups_admin(query)
    elif action == "admin_export":
        await export_users_excel(query, context)
    elif action == "admin_diagnose":
        await show_diagnose_admin(query)
    elif action == "admin_clear_queue":
        await clear_queue_admin(query)
    elif action == "admin_clear_failed":
        await clear_failed_admin(query)


async def show_admin_users(query):
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(User).order_by(User.created_at.desc()).limit(20)
        )
        users = result.scalars().all()
    
    text = f"👥 <b>Пользователи ({len(users)}):</b>\n\n"
    for u in users:
        projects_count = await get_user_projects_count(u.telegram_id)
        text += f"• {u.full_name or '—'} (@{u.username or '—'})\n"
        text += f"  🆔 {u.telegram_id} | 📁 {projects_count} проектов\n\n"
    
    keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="HTML")


async def create_backup_admin(query):
    await query.edit_message_text("📦 Создаю бэкап...")
    
    backup_service = BackupService()
    backup_path = backup_service.create_backup()
    
    if backup_path:
        try:
            with open(backup_path, 'rb') as f:
                await query.message.reply_document(
                    document=f,
                    filename=os.path.basename(backup_path),
                    caption=f"✅ Бэкап создан\n📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}"
                )
        except Exception as e:
            logger.error(f"Failed to send backup file: {e}")
        
        keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
        await query.edit_message_text(
            f"✅ Бэкап создан и отправлен!\n\n📁 {os.path.basename(backup_path)}",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    else:
        keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
        await query.edit_message_text(
            "❌ Ошибка создания бэкапа",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )


async def list_backups_admin(query):
    backup_service = BackupService()
    backups = backup_service.list_backups()
    
    if not backups:
        text = "📭 Бэкапов нет"
    else:
        text = "📦 <b>Бэкапы:</b>\n\n"
        for b in backups[:10]:
            text += f"• {b['name']}\n"
            text += f"  📅 {b['created']} | 📦 {b['size_mb']} MB\n\n"
    
    keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="HTML")


async def export_users_excel(query, context):
    await query.edit_message_text("📊 Формирую отчёт...")
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(User).order_by(User.created_at.desc()))
        users = result.scalars().all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    
    headers = ["Telegram ID", "Username", "Full Name", "Admin", "Projects", "Parsed Today", "Posted Today", "Created At"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, u in enumerate(users, 2):
        projects_count = await get_user_projects_count(u.telegram_id)
        ws.cell(row=row, column=1, value=u.telegram_id)
        ws.cell(row=row, column=2, value=u.username or "")
        ws.cell(row=row, column=3, value=u.full_name or "")
        ws.cell(row=row, column=4, value="Да" if u.is_admin else "Нет")
        ws.cell(row=row, column=5, value=projects_count)
        ws.cell(row=row, column=6, value=u.posts_parsed_today)
        ws.cell(row=row, column=7, value=u.posts_posted_today)
        ws.cell(row=row, column=8, value=u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else "")
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    await context.bot.send_document(
        chat_id=query.message.chat_id,
        document=output,
        filename=f"users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        caption="📊 Экспорт пользователей"
    )
    
    keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
    await query.edit_message_text(
        "✅ Отчёт отправлен!",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def show_diagnose_admin(query):
    text = "🔍 <b>Диагностика системы</b>\n\n"
    
    if os.path.exists(Config.DB_PATH):
        size = os.path.getsize(Config.DB_PATH) / (1024 * 1024)
        text += f"📁 БД: {Config.DB_PATH} ({size:.2f} MB)\n"
    else:
        text += f"❌ БД не найдена: {Config.DB_PATH}\n"
    
    text += f"📂 Data: {'✅' if os.path.exists(Config.DATA_DIR) else '❌'}\n"
    text += f"📂 Temp: {'✅' if os.path.exists(Config.TEMP_DIR) else '❌'}\n"
    text += f"📂 Backups: {'✅' if os.path.exists(Config.BACKUP_DIR) else '❌'}\n"
    
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(func.count()).select_from(User))
        users_count = result.scalar()
        result = await session.execute(select(func.count()).select_from(Project))
        projects_count = result.scalar()
        result = await session.execute(select(PostQueue).where(PostQueue.status == "pending"))
        pending = len(result.scalars().all())
    
    text += f"\n👥 Пользователей: {users_count}\n"
    text += f"📁 Проектов: {projects_count}\n"
    text += f"📬 В очереди: {pending}\n"
    
    keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="HTML")


async def clear_queue_admin(query):
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(PostQueue).where(PostQueue.status == "pending"))
        items = result.scalars().all()
        deleted = len(items)
        for item in items:
            await session.delete(item)
        await session.commit()
    
    keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
    await query.edit_message_text(
        f"✅ Удалено {deleted} постов из очереди",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def clear_failed_admin(query):
    async with AsyncSessionLocal() as session:
        await session.execute(delete(PostQueue).where(PostQueue.status == "failed"))
        await session.commit()
    
    keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
    await query.edit_message_text(
        "✅ Failed посты удалены",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def admin_back_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    keyboard = [
        [InlineKeyboardButton("👥 Пользователи", callback_data="admin_users_list")],
        [InlineKeyboardButton("💾 Создать бэкап", callback_data="admin_backup_create")],
        [InlineKeyboardButton("📦 Список бэкапов", callback_data="admin_backup_list")],
        [InlineKeyboardButton("📊 Экспорт в Excel", callback_data="admin_export")],
        [InlineKeyboardButton("🔍 Диагностика", callback_data="admin_diagnose")],
        [InlineKeyboardButton("🧹 Очистить очередь", callback_data="admin_clear_queue")],
        [InlineKeyboardButton("🗑️ Очистить failed", callback_data="admin_clear_failed")],
    ]
    
    await query.edit_message_text(
        "👑 <b>Админ-панель</b>\n\nВыберите действие:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="HTML"
    )